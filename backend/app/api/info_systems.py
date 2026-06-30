"""信息系统管理 API — CRUD + 导入导出。"""
import io, uuid, time
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl

from app.database import get_db
from app.models.info_system import InfoSystem, DjDjRecord, IcpRecord, SupplyChain
from app.models.user import User
from app.models.department import Department
from app.utils.security import hash_password
from app.api.deps import get_current_user, require_admin

router = APIRouter(prefix="/info-systems", tags=["信息系统"])


def _is_admin(user) -> bool:
    """判断当前用户是否为管理员。"""
    if hasattr(user, "role"):
        role = user.role
        if hasattr(role, "value"):
            role = role.value
        return role == "admin"
    return False


def _check_owner(db, model, rec_id: int, user, admin_only=True):
    """检查记录所有权：管理员可操作全部，普通用户只能操作自己创建的。"""
    rec = db.query(model).get(rec_id)
    if not rec:
        return None, "记录不存在"
    if admin_only and not _is_admin(user):
        return None, "仅管理员可操作"
    return rec, None


_asset_id_counter = 0

def _gen_asset_id(db: Session = None) -> str:
    """生成唯一资产ID，重试直到不冲突。"""
    global _asset_id_counter
    for _ in range(100):
        _asset_id_counter += 1
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        aid = f"IS-{ts}-{uuid.uuid4().hex[:8]}-{_asset_id_counter:04d}"
        if db is None:
            return aid
        if not db.query(InfoSystem).filter(InfoSystem.asset_id == aid).first():
            return aid
    raise HTTPException(500, "无法生成唯一资产ID")


# ── 人员查询 + 自动注册 ──

@router.get("/staff-search")
def search_staff(q: str = Query("", min_length=1), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """按姓名或工号搜索系统用户，未找到则返回空。"""
    items = db.query(User).filter(
        (User.name.contains(q)) | (User.gh.contains(q)) | (User.username.contains(q))
    ).limit(15).all()
    return {"items": [{"id": u.id, "name": u.name, "gh": u.gh, "username": u.username,
                        "phone": u.phone, "mobile": u.mobile, "department_name": u.department.dwmc if u.department else ""} for u in items]}


@router.get("/staff-lookup")
def lookup_staff(q: str = Query("", min_length=1), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """查找人员（系统用户 + 数据中台教职工API），返回合并结果供确认。"""
    results = []
    # 1. 系统用户
    users = db.query(User).filter(
        (User.name.contains(q)) | (User.gh.contains(q)) | (User.username.contains(q))
    ).limit(10).all()
    existing_ghs = {u.gh for u in users if u.gh}
    for u in users:
        results.append({"id": u.id, "name": u.name, "gh": u.gh, "username": u.username,
                        "phone": u.phone, "mobile": u.mobile,
                        "department_name": u.department.dwmc if u.department else "",
                        "source": "系统用户"})
    # 2. 本地教职工库（排除已有用户）
    from app.models.staff_info import StaffInfo
    staffs = db.query(StaffInfo).filter(
        (StaffInfo.xm.contains(q)) | (StaffInfo.gh.contains(q))
    ).limit(10).all()
    for s in staffs:
        if s.gh and s.gh in existing_ghs:
            continue
        results.append({"id": None, "name": s.xm, "gh": s.gh, "username": None,
                        "phone": s.bgdh or "", "mobile": s.yddh or "",
                        "department_name": s.szdwbm or "",
                        "source": "教职工库（可注册）"})
    # 3. 数据中台教职工API查询（排除已有用户和已查到的）
    seen_ghs = existing_ghs | {s.gh for s in staffs if s.gh}
    try:
        from app.services.external_api_service import fetch_staff
        # 尝试按工号和姓名分别查询
        api_results = []
        try:
            api_results.extend(fetch_staff(db, gh=q))
        except Exception:
            pass
        try:
            api_results.extend(fetch_staff(db, xm=q))
        except Exception:
            pass
        # 去重
        api_seen = set()
        for r in api_results:
            gh = (r.get("GH") or "").strip()
            if not gh or gh in seen_ghs or gh in api_seen:
                continue
            api_seen.add(gh)
            xm = (r.get("XM") or "").strip()
            email = (r.get("DZYX") or "").strip()
            phone = (r.get("BGDH") or "").strip()
            mobile = (r.get("YDDH") or "").strip()
            szdwbm = (r.get("SZDWBM") or "").strip()
            # 尝试匹配部门名称
            dept_name = szdwbm
            if szdwbm:
                d = db.query(Department).filter(Department.dwbm == szdwbm).first()
                if d:
                    dept_name = d.dwmc
            results.append({"id": None, "name": xm, "gh": gh, "username": None,
                            "phone": phone, "mobile": mobile, "email": email,
                            "department_name": dept_name,
                            "source": "数据中台（可注册）"})
        seen_ghs |= api_seen
    except Exception:
        pass  # API不可用时不影响本地查询
    return {"items": results[:20]}


@router.post("/staff-register")
def register_staff(body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """查询教职工并自动注册为系统用户（如不存在）。返回用户信息。"""
    name = (body.get("name") or "").strip()
    gh = (body.get("gh") or "").strip()
    if not name and not gh:
        raise HTTPException(400, "请提供姓名或工号")
    # 先在本地查找
    q = db.query(User)
    if name and gh:
        q = q.filter(User.name == name, User.gh == gh)
    elif gh:
        q = q.filter(User.gh == gh)
    elif name:
        q = q.filter(User.name == name)
    user = q.first()
    if user:
        return {"id": user.id, "name": user.name, "gh": user.gh, "username": user.username,
                "phone": user.phone, "mobile": user.mobile, "department_name": user.department.dwmc if user.department else "",
                "existed": True}

    def _find_or_create_user(staff_data: dict) -> dict:
        """根据API/本地数据查找或创建用户。"""
        s_gh = (staff_data.get("gh") or "").strip()
        s_name = (staff_data.get("name") or "").strip()
        s_email = (staff_data.get("email") or "").strip() or None
        s_phone = (staff_data.get("phone") or "").strip() or None
        s_mobile = (staff_data.get("mobile") or "").strip() or None
        s_gender = (staff_data.get("gender") or "").strip() or ""
        s_dept_code = (staff_data.get("dept_code") or "").strip()
        # 匹配部门
        dept_id = None
        if s_dept_code:
            d = db.query(Department).filter(Department.dwbm == s_dept_code).first()
            if d:
                dept_id = d.id
        username = s_gh or f"u{int(time.time())}"
        exist = db.query(User).filter(User.username == username).first()
        if exist:
            username = f"{s_gh}_{int(time.time() % 1000)}"
        new_user = User(
            username=username,
            password_hash=hash_password("Abcd1234!"),
            name=s_name,
            gh=s_gh,
            gender=s_gender,
            phone=s_phone,
            mobile=s_mobile,
            email=s_email,
            department_id=dept_id,
            user_type="internal",
            role="user",
            is_active=True,
        )
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        return {"id": new_user.id, "name": new_user.name, "gh": new_user.gh, "username": new_user.username,
                "phone": new_user.phone, "mobile": new_user.mobile,
                "department_name": new_user.department.dwmc if new_user.department else "",
                "existed": False, "message": f"已自动注册: {new_user.name}"}

    # 1. 尝试从数据中台API查找
    try:
        from app.services.external_api_service import fetch_staff
        api_results = []
        if gh:
            api_results = fetch_staff(db, gh=gh)
        if not api_results and name:
            api_results = fetch_staff(db, xm=name)
        if api_results:
            r = api_results[0]
            return _find_or_create_user({
                "gh": (r.get("GH") or "").strip(),
                "name": (r.get("XM") or "").strip(),
                "email": (r.get("DZYX") or "").strip(),
                "phone": (r.get("BGDH") or "").strip(),
                "mobile": (r.get("YDDH") or "").strip(),
                "gender": "男" if r.get("XBM") == "1" else ("女" if r.get("XBM") == "2" else ""),
                "dept_code": (r.get("SZDWBM") or "").strip(),
            })
    except Exception:
        pass  # API不可用则降级到本地

    # 2. 尝试从本地 staff_info 查找
    from app.models.staff_info import StaffInfo
    sq = db.query(StaffInfo)
    if name and gh:
        sq = sq.filter(StaffInfo.xm == name, StaffInfo.gh == gh)
    elif gh:
        sq = sq.filter(StaffInfo.gh == gh)
    elif name:
        sq = sq.filter(StaffInfo.xm == name)
    staff = sq.first()
    if staff:
        return _find_or_create_user({
            "gh": staff.gh or "",
            "name": staff.xm,
            "email": staff.dzyx or "",
            "phone": staff.bgdh or "",
            "mobile": staff.yddh or "",
            "gender": "男" if staff.xbm == "1" else ("女" if staff.xbm == "2" else ""),
            "dept_code": staff.szdwbm or "",
        })
    raise HTTPException(404, "未在系统中找到匹配的教职工信息，请确认姓名或工号")


# ── InfoSystem CRUD ──

_IS_SORTABLE = {"system_name","system_type","sub_type","ip_address","domain","entry_url",
    "manager_name","owner_name","fill_type","url_status","remark","belong_dept_name"}

@router.get("")
def list_systems(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
                 search: str = Query(""), fill_type: str = Query(""),
                 system_type: str = Query(""), sub_type: str = Query(""),
                 manager_name: str = Query(""), owner_name: str = Query(""),
                 url_status: str = Query(""),
                 sort_field: str = Query(""), sort_order: str = Query("desc"),
                 db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(InfoSystem)
    # 非管理员：本单位 + (自己是管理员 或 未分配管理员)
    if not _is_admin(user):
        uid = str(user.gh or user.id)
        dept_id = getattr(user, 'department_id', None)
        if dept_id:
            q = q.filter((InfoSystem.dept_id == dept_id) | (InfoSystem.dept_id == None))
        q = q.filter((InfoSystem.manager_gh == uid) | (InfoSystem.manager_gh == None) | (InfoSystem.manager_gh == ""))
    if search:
        kw = f"%{search}%"
        q = q.filter(InfoSystem.system_name.like(kw) | InfoSystem.ip_address.like(kw) | InfoSystem.domain.like(kw) | InfoSystem.manager_name.like(kw) | InfoSystem.owner_name.like(kw))
    if fill_type:
        q = q.filter(InfoSystem.fill_type == fill_type)
    if system_type:
        q = q.filter(InfoSystem.system_type == system_type)
    if sub_type:
        q = q.filter(InfoSystem.sub_type == sub_type)
    if manager_name:
        q = q.filter(InfoSystem.manager_name.like(f"%{manager_name}%"))
    if owner_name:
        q = q.filter(InfoSystem.owner_name.like(f"%{owner_name}%"))
    if url_status:
        q = q.filter(InfoSystem.url_status == url_status)
    total = q.count()
    # 排序：白名单字段 + 方向
    if sort_field and sort_field in _IS_SORTABLE:
        col = getattr(InfoSystem, sort_field, None)
        if col is not None:
            q = q.order_by(col.desc()) if sort_order == "desc" else q.order_by(col.asc())
        else:
            q = q.order_by(InfoSystem.id.desc())
    else:
        q = q.order_by(InfoSystem.id.desc())
    items = q.offset((page - 1) * size).limit(size).all()
    # 批量查询部门名称
    dept_ids = {r.dept_id for r in items if r.dept_id}
    dept_map = {}
    if dept_ids:
        depts = db.query(Department).filter(Department.id.in_(dept_ids)).all()
        dept_map = {d.id: d.dwmc for d in depts}
    result = []
    for r in items:
        d = {c.name: getattr(r, c.name) for c in r.__table__.columns}
        d["belong_dept_name"] = dept_map.get(r.dept_id, "")
        result.append(d)
    return {"items": result, "total": total}


# 批量认领/撤销
@router.post("/batch-claim")
def batch_claim(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    models = {"info_system": InfoSystem, "djdj": DjDjRecord, "icp": IcpRecord, "supply_chain": SupplyChain}
    m = models.get(body.get("model", ""))
    if not m: raise HTTPException(400, "不支持的类型")
    from datetime import datetime, timezone, timedelta
    tz = timezone(timedelta(hours=8))
    now = datetime.now(tz).replace(tzinfo=None)
    count = 0
    for rid in body.get("ids", []):
        rec = db.query(m).get(rid)
        if not rec: continue
        if body.get("model") == "info_system":
            # 信息系统：认领 = 设置为管理员
            if rec.manager_gh and rec.manager_gh.strip():
                continue  # 已有管理员
            rec.manager_gh = str(user.gh or user.id)
            rec.manager_name = user.name or user.username
            if not rec.created_by:
                rec.created_by = user.id
            count += 1
        elif rec.claimed_by is None:
            rec.claimed_by = user.id
            rec.claimed_at = now
            if not rec.created_by:
                rec.created_by = user.id
            count += 1
    db.commit()
    return {"message": f"成功认领 {count} 条", "count": count}


@router.post("/batch-cancel")
def batch_cancel_systems(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """批量申请注销（本人认领的记录可申请注销）。"""
    models = {"info_system": InfoSystem, "djdj": DjDjRecord, "icp": IcpRecord, "supply_chain": SupplyChain}
    m = models.get(body.get("model", ""))
    if not m: raise HTTPException(400, "不支持的类型")
    ids = body.get("ids", [])
    if not ids: raise HTTPException(400, "请选择记录")
    user_gh = str(user.gh or user.id)
    count = 0
    for rid in ids:
        rec = db.query(m).get(rid)
        if not rec: continue
        # 只有管理员或本人认领的才能申请注销
        is_owner = False
        if hasattr(rec, 'manager_gh') and str(rec.manager_gh or '') == user_gh:
            is_owner = True
        elif hasattr(rec, 'claimed_by') and rec.claimed_by == user.id:
            is_owner = True
        elif hasattr(rec, 'created_by') and rec.created_by == user.id:
            is_owner = True
        if _is_admin(user) or is_owner:
            if hasattr(rec, 'fill_type'):
                rec.fill_type = "申请注销"
            count += 1
    db.commit()
    return {"message": f"已申请注销 {count} 条"}


@router.post("/batch-revoke")
def batch_revoke(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    models = {"info_system": InfoSystem, "djdj": DjDjRecord, "icp": IcpRecord, "supply_chain": SupplyChain}
    m = models.get(body.get("model", ""))
    if not m: raise HTTPException(400, "不支持的类型")
    count = 0
    for rid in body.get("ids", []):
        rec = db.query(m).get(rid)
        if not rec: continue
        if body.get("model") == "info_system":
            # 信息系统：撤销 = 清空管理员
            if rec.manager_gh and str(rec.manager_gh) == str(user.gh or user.id):
                rec.manager_gh = ""
                rec.manager_name = ""
                count += 1
        elif rec and (rec.claimed_by == user.id or _is_admin(user)):
            rec.claimed_by = None
            rec.claimed_at = None
            count += 1
    db.commit()
    return {"message": f"成功撤销 {count} 条", "count": count}


# 认领/撤销认领（通用）
@router.post("/claim/{model}/{rec_id}")
def claim_record(model: str, rec_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """认领一条记录（信息系统/等保/ICP/供应链）。"""
    models = {"info_system": InfoSystem, "djdj": DjDjRecord, "icp": IcpRecord, "supply_chain": SupplyChain}
    m = models.get(model)
    if not m:
        raise HTTPException(400, "不支持的类型")
    rec = db.query(m).get(rec_id)
    if not rec:
        raise HTTPException(404, "记录不存在")
    if rec.claimed_by is not None:
        raise HTTPException(400, "该记录已被认领")
    from datetime import datetime, timezone, timedelta
    tz = timezone(timedelta(hours=8))
    rec.claimed_by = user.id
    rec.claimed_at = datetime.now(tz).replace(tzinfo=None)
    if not rec.created_by:
        rec.created_by = user.id
    db.commit()
    return {"message": "认领成功"}


@router.post("/revoke/{model}/{rec_id}")
def revoke_record(model: str, rec_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """撤销认领。"""
    models = {"info_system": InfoSystem, "djdj": DjDjRecord, "icp": IcpRecord, "supply_chain": SupplyChain}
    m = models.get(model)
    if not m:
        raise HTTPException(400, "不支持的类型")
    rec = db.query(m).get(rec_id)
    if not rec:
        raise HTTPException(404, "记录不存在")
    if rec.claimed_by != user.id and not _is_admin(user):
        raise HTTPException(403, "只能撤销自己的认领")
    rec.claimed_by = None
    rec.claimed_at = None
    db.commit()
    return {"message": "已撤销认领"}


@router.post("/batch-delete")
def batch_delete_systems(body: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "请选择记录")
    db.query(InfoSystem).filter(InfoSystem.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {len(ids)} 条"}


# 不可由前端直接修改的只读列
def _ensure_supply_chain(db, vendor_name: str, created_by: int = None):
    """如果开发厂商名称不在 supply_chains 表中，则自动创建。"""
    if not vendor_name or not vendor_name.strip():
        return
    name = vendor_name.strip()
    exist = db.query(SupplyChain).filter(SupplyChain.company_name == name).first()
    if not exist:
        try:
            db.add(SupplyChain(company_name=name, created_by=created_by))
            db.commit()
        except Exception:
            db.rollback()

_READONLY_COLS = {"id", "created_at", "updated_at"}

def _sanitize_body(data: dict, model_class=None) -> dict:
    """将非字符串列的空字符串转为 None，避免 MySQL 数据类型错误。

    对 Date/DateTime/Integer/Float/Numeric 等列，空字符串 "" → None。
    对 String/Text 列保留空字符串（合法值）。
    始终排除 id/created_at/updated_at 只读列。
    """
    from sqlalchemy import Date, DateTime, Integer, Float, Numeric, BigInteger, SmallInteger
    if model_class is None:
        model_class = InfoSystem
    non_str_types = (Date, DateTime, Integer, Float, Numeric, BigInteger, SmallInteger)
    # 获取需要特殊处理的列名
    cols_needing_none = set()
    for c in model_class.__table__.columns:
        if isinstance(c.type, non_str_types):
            cols_needing_none.add(c.name)
    # 也加入 Date 列（来自旧版兼容）
    result = {}
    for k, v in data.items():
        if k in _READONLY_COLS:
            continue  # 静默排除只读列
        if k in cols_needing_none and v == "":
            result[k] = None
        else:
            result[k] = v
    return result


@router.post("")
def create_system(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not body.get("asset_id"):
        body["asset_id"] = _gen_asset_id(db)
    if db.query(InfoSystem).filter(InfoSystem.asset_id == body["asset_id"]).first():
        raise HTTPException(400, "资产ID已存在")
    clean = _sanitize_body(body, InfoSystem)
    if not clean.get("system_name"):
        raise HTTPException(400, "系统名称不能为空")
    clean["created_by"] = user.id
    sys = InfoSystem(**clean)
    try:
        db.add(sys); db.commit(); db.refresh(sys)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"创建失败：{e}")
    # 自动创建供应链记录
    _ensure_supply_chain(db, clean.get("vendor_name", ""), user.id)
    return {"id": sys.id, "message": "已创建"}


@router.put("/{sys_id}")
def update_system(sys_id: int, body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    sys = db.query(InfoSystem).get(sys_id)
    if not sys:
        raise HTTPException(404, "记录不存在")
    # 非管理员只能编辑自己创建的数据
    # 管理员/创建人可编辑，负责人仅浏览
    user_gh = str(user.gh or user.id)
    if not _is_admin(user) and sys.created_by != user.id and user_gh != str(sys.manager_gh or ""):
        if sys.owner_gh and user_gh == str(sys.owner_gh):
            raise HTTPException(403, "负责人只能浏览，不能编辑数据")
        raise HTTPException(403, "只能编辑自己创建或管理的数据")
    clean = _sanitize_body(body, InfoSystem)
    try:
        for k, v in clean.items():
            setattr(sys, k, v)
        db.commit()
        print(f"[UPDATE] sys_id={sys_id} 更新成功, 接收字段数={len(clean)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"保存失败：{e}")
    # 自动创建供应链记录
    _ensure_supply_chain(db, clean.get("vendor_name", ""), user.id)
    return {"message": "已更新"}


@router.delete("/{sys_id}")
def delete_system(sys_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    sys = db.query(InfoSystem).get(sys_id)
    if not sys:
        raise HTTPException(404, "记录不存在")
    if not _is_admin(user) and sys.created_by != user.id:
        raise HTTPException(403, "只能删除自己创建的数据")
    db.delete(sys)
    db.commit()
    return {"message": "已删除"}


# ── Excel 导入 ──

# ── 导入格式定义 ──

def _detect_import_format(ws) -> str:
    """检测 Excel 文件格式，返回格式名称。"""
    # 读取前两行的列值
    row1 = [_cell_str(ws.cell(1, c).value) for c in range(1, min(ws.max_column + 1, 80))]
    row2 = [_cell_str(ws.cell(2, c).value) for c in range(1, min(ws.max_column + 1, 80))]
    r1 = " ".join(row1)
    r2 = " ".join(row2)
    # 格式1: 系统导出 info_systems(*).xlsx（行1含"资产ID"+"系统名称"，排在iso_report之前）
    if "资产ID" in row1 and "系统名称" in row1 and "资产类型" in row1:
        return "system_export"
    # 格式2: 信息系统管理_*.xlsx（行1是中文表头，含"单位名称"+"管理员姓名"）
    if "单位名称" in r1 and "管理员姓名" in r1:
        return "iso_report"
    # 格式3: 资产清单.xlsx（行1是说明文字，行2是"ID"+"所属单位"+"信息系统名称"等表头）
    if ("说明" in r1 or "必填" in r1 or len(row1[0]) > 50) and "ID" in row2 and "所属单位" in r2:
        return "asset_list"
    # 格式4: 资产导出.xlsx（行1是长说明，行2直接是数据或无表头行）
    if "填写说明" in r1 or "必填项" in r1 or len(row1[0]) > 50:
        return "asset_export"
    return "unknown"


# 旧格式（资产导出.xlsx）列映射：列号 → (模型字段, 转换函数或None)
_FORMAT_ASSET_EXPORT = {
    3:  "system_name",
    4:  "system_type",
    5:  "sub_type",
    6:  "ip_address",
    7:  "domain",
    8:  "djdj_status",
    12: "djdj_sys_name",
    13: "djdj_date",
    14: "djdj_no",
    2:  "org_name",
    25: "dept_name",
    27: "contact",
    28: "contact_phone",
    11: "remark",
}

# 新格式（信息系统管理_*.xlsx）列映射：列号 → (模型字段, 转换函数或None)
_FORMAT_ISO_REPORT = {
    3:  "org_name",           # 单位名称
    4:  "_dept_code",         # 单位名称_代码 → 查 departments.dwbm 得到 dept_id
    5:  "manager_name",       # 管理员姓名
    6:  "manager_gh",         # 管理员工号_账号 → 无账号则自动创建
    8:  "_manager_mobile",    # 管理员手机号码 → 创建账号时填入 mobile
    10: "system_type",        # 系统类型
    11: "sub_type",           # 信息系统类型 (K列，主键)
    12: "product_name",       # 产品名称
    13: "product_version",    # 版本号
    14: "domain",             # 域名
    15: "ip_address",         # 业务互联网IP
    16: "ip_address2",        # 业务内网IP（合并到ip_address）
    18: "dept_name",          # 运维单位
    24: "source_type",        # 数据来源
    41: "djdj_status",        # 等保定级情况
    42: "djdj_sys_name",      # 等保二级系统名称 → 等保信息维护表联动
    43: "djdj_no",            # 等保定级编号
    44: "djdj_date",          # 等保定级时间
    47: "icp_no",             # ICP备案号
    49: "icp_date",           # ICP备案时间
    51: "remark",             # 备注
    52: "vendor_name",        # 供应链单位名称 → 供应链信息维护表联动
    56: "vendor_contact",     # 供应链联系人
    57: "vendor_phone",       # 供应链联系电话
    70: "submitter",          # 提交人
}


def _parse_excel_date(val):
    """尝试将值解析为日期字符串。"""
    if val is None: return None
    if isinstance(val, (datetime,)): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s: return None
    # 尝试常见日期格式
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            datetime.strptime(s[:19] if len(s) > 19 else s, fmt)
            return s[:10]
        except ValueError:
            continue
    return s[:10] if len(s) >= 10 else None


def _is_empty(val) -> bool:
    """判断值是否为空。"""
    if val is None: return True
    if isinstance(val, str) and not val.strip(): return True
    return False


def _parse_row_asset_export(ws, row_idx: int) -> dict:
    """解析旧格式的一行数据。"""
    data = {"fill_type": "导入"}
    for col, field in _FORMAT_ASSET_EXPORT.items():
        raw = ws.cell(row_idx, col).value
        if field == "djdj_date":
            data[field] = _parse_excel_date(raw)
        else:
            data[field] = _cell_str(raw)
    # 合并内网IP（如果有第16列）
    ip2 = _cell_str(ws.cell(row_idx, 16).value)
    if ip2 and data.get("ip_address"):
        data["ip_address"] = data["ip_address"] + ("," if not data["ip_address"].endswith(",") else "") + ip2
    elif ip2:
        data["ip_address"] = ip2
    return data


# 资产清单.xlsx 格式：表头名 → 模型字段
_FORMAT_ASSET_LIST_HEADER_MAP = {
    "所属单位":              "org_name",
    "信息系统名称":          "system_name",
    "资产类型":              "system_type",
    "信息系统类型":          "sub_type",
    "IP":                   "ip_address",
    "域名":                 "domain",
    "等保备案级别":          "djdj_level",
    "备注":                 "remark",
    "等保备案系统名称":      "djdj_sys_name",
    "等保备案时间":          "djdj_date",
    "等保备案编号":          "djdj_no",
    "联系人":               "contact",
    "联系人电话":           "contact_phone",
    "系统运维单位":          "dept_name",
    "内网URL":              "internal_url",
    "内网IP地址段（只明确到C段）": "ip_range",
    # 供应链第1组
    "供应链厂商情况-开发厂商名称-1": "vendor_name",
    "供应链厂商情况-产品名称-1":   "product_name",
    "供应链厂商情况-版本号-1":     "product_version",
    "供应链厂商情况-来源-1":       "source_type",
    # 供应链第2组
    "供应链厂商情况-开发厂商名称-2": "_vendor2_name",
    "供应链厂商情况-产品名称-2":   "_vendor2_product",
    "供应链厂商情况-版本号-2":     "_vendor2_version",
    "供应链厂商情况-来源-2":       "_vendor2_source",
}
# 资产清单中需要合并到 ip_address 的额外列
_ASSET_LIST_EXTRA_IP_HEADERS = ["互联网接入IP地址", "IP"]


# 系统导出 info_systems(*).xlsx 格式：中文表头 → 模型字段
_SYSTEM_EXPORT_HEADER_MAP = {
    "资产ID":          "asset_id",
    "系统名称":        "system_name",
    "资产类型":        "system_type",
    "信息系统类型":    "sub_type",
    "IP地址":         "ip_address",
    "域名/URL":       "domain",
    "单位名称":        "org_name",
    "运维单位":        "dept_name",
    "联系人":         "contact",
    "联系电话":       "contact_phone",
    "填报类型":       "fill_type",
    "所属部门ID":     "dept_id",
    "管理员姓名":     "manager_name",
    "管理员工号":     "manager_gh",
    "负责人姓名":     "owner_name",
    "负责人工号":     "owner_gh",
    "等保状态":       "djdj_status",
    "等保系统名称":   "djdj_sys_name",
    "等保编号":       "djdj_no",
    "等保等级":       "djdj_level",
    "等保日期":       "djdj_date",
    "测评单位":       "djdj_org",
    "ICP备案号":      "icp_no",
    "ICP备案日期":    "icp_date",
    "开发厂商":       "vendor_name",
    "产品名称":       "product_name",
    "版本号":         "product_version",
    "来源":           "source_type",
    "厂商联系人":     "vendor_contact",
    "厂商电话":       "vendor_phone",
    "运维联系人":     "ops_contact",
    "运维电话":       "ops_phone",
    "是否有网站":     "has_website",
    "内网URL":        "internal_url",
    "IP地址段":       "ip_range",
    "入口地址":       "entry_url",
    "验证状态":       "url_status",
    "备注":           "remark",
}
_SYSTEM_EXPORT_INT_FIELDS = {"dept_id"}
_SYSTEM_EXPORT_DATE_FIELDS = {"djdj_date", "icp_date"}


def _build_system_export_col_map(ws) -> dict:
    """从第1行表头构建 列号→模型字段 的映射。"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        header = _cell_str(ws.cell(1, c).value).strip()
        if header in _SYSTEM_EXPORT_HEADER_MAP:
            col_map[c] = _SYSTEM_EXPORT_HEADER_MAP[header]
    return col_map


def _parse_row_system_export(ws, row_idx: int, col_map: dict) -> dict:
    """解析系统导出格式的一行数据。"""
    data = {}
    for c, field in col_map.items():
        raw = ws.cell(row_idx, c).value
        if field in _SYSTEM_EXPORT_DATE_FIELDS:
            data[field] = _parse_excel_date(raw)
        elif field in _SYSTEM_EXPORT_INT_FIELDS:
            try:
                data[field] = int(raw) if raw is not None and str(raw).strip() else None
            except (ValueError, TypeError):
                data[field] = None
        else:
            data[field] = _cell_str(raw)
    return data


def _build_asset_list_col_map(ws) -> dict:
    """从第2行表头构建 列号→模型字段 的映射。"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        header = _cell_str(ws.cell(2, c).value).strip()
        if header in _FORMAT_ASSET_LIST_HEADER_MAP:
            col_map[c] = _FORMAT_ASSET_LIST_HEADER_MAP[header]
    return col_map


def _parse_row_asset_list(ws, row_idx: int, col_map: dict) -> dict:
    """解析资产清单.xlsx 的一行数据。"""
    data = {"fill_type": "导入"}
    ip_parts = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row_idx, c).value
        val_str = _cell_str(raw)
        header = _cell_str(ws.cell(2, c).value).strip()
        # 收集所有 IP 列（主IP + 内部IP等）
        if header in _ASSET_LIST_EXTRA_IP_HEADERS and val_str:
            ip_parts.append(val_str)
        if c in col_map:
            field = col_map[c]
            if field == "djdj_date":
                data[field] = _parse_excel_date(raw)
            else:
                data[field] = val_str
    # 合并 IP
    if ip_parts:
        data["ip_address"] = ",".join(dict.fromkeys(ip_parts))  # 去重保序
    return data


def _parse_row_iso_report(ws, row_idx: int) -> dict:
    """解析新格式的一行数据。"""
    data = {"fill_type": "导入"}
    for col, field in _FORMAT_ISO_REPORT.items():
        raw = ws.cell(row_idx, col).value
        if field in ("djdj_date", "icp_date"):
            data[field] = _parse_excel_date(raw)
        elif field == "ip_address2":
            data["_ip_internal"] = _cell_str(raw)
        elif field == "_dept_code":
            data["_dept_code"] = _cell_str(raw)
        elif field == "_manager_mobile":
            data["_manager_mobile"] = _cell_str(raw)
        elif field == "submitter":
            data["contact"] = _cell_str(raw)
        else:
            data[field] = _cell_str(raw)
    # 提取系统名称（主键：K列=信息系统类型 → 产品名称 → 系统类型）
    system_name = data.get("sub_type", "") or data.get("product_name", "") or data.get("system_type", "")
    data["system_name"] = system_name
    # 合并IP
    ip_ext = data.get("ip_address", "")
    ip_int = data.pop("_ip_internal", "")
    if ip_int:
        data["ip_address"] = (ip_ext + ("," if ip_ext and not ip_ext.endswith(",") else "") + ip_int) if ip_ext else ip_int
    return data


# ── 导入端点 ──

@router.post("/import")
def import_excel(file: UploadFile, mode: str = Query("supplement"),
                 db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    导入 Excel 文件，自动识别格式。
    mode: overwrite=覆盖已有 / supplement=补充空白字段 / skip=跳过重复
    """
    if mode not in ("overwrite", "supplement", "skip"):
        raise HTTPException(400, "mode 必须是 overwrite / supplement / skip 之一")
    try:
        return _do_import(file, mode, db)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(500, f"导入失败：{e}\n{traceback.format_exc()}")


def _do_import(file: UploadFile, mode: str, db: Session):
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active

    # 检测格式
    fmt = _detect_import_format(ws)
    if fmt == "unknown":
        raise HTTPException(400, "无法识别文件格式，请确认文件类型。支持：资产导出.xlsx、信息系统管理_*.xlsx")

    # 找到数据起始行
    if fmt == "system_export":
        start_row = 2  # 行1=表头, 行2起=数据
        col_map = _build_system_export_col_map(ws)
        parse_row = lambda ws, ri: _parse_row_system_export(ws, ri, col_map)
        fmt_name = "系统导出"
        # 系统导出按 asset_id 去重（保留原ID）
        dedup_by = "asset_id"
    elif fmt == "asset_export":
        start_row = _find_data_start(ws)
        parse_row = _parse_row_asset_export
        fmt_name = "资产导出"
        dedup_by = "system_name"
    elif fmt == "asset_list":
        start_row = 3  # 行1=说明, 行2=表头, 行3起=数据
        col_map = _build_asset_list_col_map(ws)
        parse_row = lambda ws, ri: _parse_row_asset_list(ws, ri, col_map)
        fmt_name = "资产清单"
        dedup_by = "system_name"
    else:
        start_row = 2  # 第1行是表头，第2行开始是数据
        parse_row = _parse_row_iso_report
        fmt_name = "信息系统管理"
        dedup_by = "system_name"

    # 解析所有行
    parsed = []
    seen_keys = set()
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = parse_row(ws, row_idx)
        if dedup_by == "asset_id":
            key = row_data.get("asset_id", "").strip()
        else:
            key = row_data.get("system_name", "").strip()
        if not key:
            continue
        if key in seen_keys:
            continue
        seen_keys.add(key)
        parsed.append(row_data)

    # 统计 + 导入
    stats = {"total": len(parsed), "created": 0, "updated": 0, "supplemented": 0, "skipped": 0, "errors": 0,
             "dept_matched": 0, "users_created": 0, "supply_added": 0, "format": fmt_name}
    # 预加载已有供应链公司名
    existing_companies = {c[0] for c in db.query(SupplyChain.company_name).all()}
    valid_cols = {c.name for c in InfoSystem.__table__.columns}
    date_cols = {"djdj_date", "icp_date"}

    # 清理存量数据中前后带空格的系统名称
    try:
        from sqlalchemy import func as sa_func
        dirty = db.query(InfoSystem).filter(
            sa_func.trim(InfoSystem.system_name) != InfoSystem.system_name
        ).all()
        for d in dirty:
            d.system_name = d.system_name.strip()
        if dirty:
            db.flush()
    except Exception:
        pass

    # 预加载部门代码映射（仅新格式需要）
    dept_code_map = {}
    if fmt == "iso_report":
        depts = db.query(Department).filter(Department.dwbm.isnot(None)).all()
        dept_code_map = {d.dwbm: d.id for d in depts if d.dwbm}

    # 预加载已有的用户工号
    existing_ghs = {u.gh for u in db.query(User).filter(User.gh.isnot(None), User.gh != "").all()}

    # 收集需要创建账号的管理员工号 → 通过外部 API 获取详细信息
    staff_cache = {}  # gh → staff_dict (来自 fetch_staff API)
    if fmt == "iso_report":
        new_manager_ghs = []
        for row_data in parsed:
            gh = row_data.get("manager_gh", "").strip()
            name = row_data.get("manager_name", "").strip()
            if gh and name and gh not in existing_ghs:
                new_manager_ghs.append(gh)
        new_manager_ghs = list(dict.fromkeys(new_manager_ghs))  # 去重保序

        if new_manager_ghs:
            try:
                from app.services.external_api_service import fetch_staff
                for gh in new_manager_ghs:
                    try:
                        results = fetch_staff(db, gh=gh)
                        if results:
                            # 优先取工号完全匹配的第一条
                            staff_cache[gh] = results[0]
                    except Exception:
                        pass  # API 不可用时跳过
            except Exception:
                pass  # 外部 API 模块不可用时跳过

    for row_data in parsed:
        try:
            system_name = row_data["system_name"].strip()
            row_data["system_name"] = system_name  # 确保写入DB的名称已去空格

            # 解析部门代码 → dept_id
            dept_code = row_data.pop("_dept_code", "")
            if dept_code and dept_code in dept_code_map:
                row_data["dept_id"] = dept_code_map[dept_code]
                stats["dept_matched"] += 1
            # 清理内部字段
            manager_mobile = row_data.pop("_manager_mobile", "")

            # 管理员自动创建账号（参考 asset_match.py 的 fetch_staff 逻辑）
            manager_gh = row_data.get("manager_gh", "").strip()
            manager_name = row_data.get("manager_name", "").strip()
            if manager_gh and manager_name and manager_gh not in existing_ghs:
                staff = staff_cache.get(manager_gh, {})
                staff_dept_code = (staff.get("SZDWBM") or "").strip()
                staff_dept_id = dept_code_map.get(staff_dept_code) if staff_dept_code else None
                email = (staff.get("DZYX") or "").strip() or None
                phone = (staff.get("BGDH") or "").strip() or None
                mobile = manager_mobile or (staff.get("YDDH") or "").strip() or None
                gender = "男" if staff.get("XBM") == "1" else "女" if staff.get("XBM") == "2" else ""
                # 检查 email 是否已被其他用户占用
                if email:
                    email_taken = db.query(User).filter(User.email == email, User.gh != manager_gh).first()
                    if email_taken:
                        email = None
                try:
                    new_user = User(
                        username=manager_gh,
                        gh=manager_gh,
                        name=manager_name,
                        user_type="internal",
                        password_hash=hash_password("Abcd1234!"),
                        email=email,
                        phone=phone,
                        mobile=mobile,
                        gender=gender,
                        department_id=staff_dept_id or (dept_code_map.get(dept_code) if dept_code else None),
                    )
                    db.add(new_user)
                    db.flush()
                    existing_ghs.add(manager_gh)
                    stats["users_created"] += 1
                except Exception:
                    db.rollback()  # 重置会话状态
                    # 重试：不带 email（可能是其它约束冲突）
                    try:
                        new_user2 = User(
                            username=manager_gh,
                            gh=manager_gh,
                            name=manager_name,
                            user_type="internal",
                            password_hash=hash_password("Abcd1234!"),
                            phone=phone,
                            mobile=mobile,
                            gender=gender,
                            department_id=staff_dept_id or (dept_code_map.get(dept_code) if dept_code else None),
                        )
                        db.add(new_user2)
                        db.flush()
                        existing_ghs.add(manager_gh)
                        stats["users_created"] += 1
                    except Exception:
                        db.rollback()

            # 匹配已有记录：系统导出按 asset_id，其他按 system_name
            from sqlalchemy import func as sa_func
            if dedup_by == "asset_id":
                asset_id = row_data.get("asset_id", "").strip()
                existing = db.query(InfoSystem).filter(InfoSystem.asset_id == asset_id).first() if asset_id else None
            else:
                existing = db.query(InfoSystem).filter(
                    sa_func.trim(InfoSystem.system_name) == system_name
                ).first()

            if existing:
                if mode == "skip":
                    stats["skipped"] += 1
                    continue
                elif mode == "overwrite":
                    for k, v in row_data.items():
                        if k != "id" and k in valid_cols:
                            if k in date_cols and _is_empty(v):
                                setattr(existing, k, None)
                            else:
                                setattr(existing, k, v)
                    existing.fill_type = "导入"
                    stats["updated"] += 1
                elif mode == "supplement":
                    changed = False
                    for k, v in row_data.items():
                        if k != "id" and k in valid_cols:
                            current = getattr(existing, k, None)
                            if _is_empty(current) and not _is_empty(v):
                                if k in date_cols:
                                    setattr(existing, k, _parse_excel_date(v))
                                else:
                                    setattr(existing, k, v)
                                changed = True
                    if changed:
                        stats["supplemented"] += 1
                    else:
                        stats["skipped"] += 1
            else:
                clean = {}
                for k, v in row_data.items():
                    if k in valid_cols:
                        if k in date_cols and _is_empty(v):
                            clean[k] = None
                        else:
                            clean[k] = v
                if not clean.get("asset_id"):
                    clean["asset_id"] = _gen_asset_id(db)
                sys_obj = InfoSystem(**clean)
                db.add(sys_obj)
                stats["created"] += 1

            # 供应链自动入库：开发厂商名称不在 supply_chains 表中则自动添加
            vendor = (row_data.get("vendor_name") or "").strip()
            if vendor and vendor not in existing_companies:
                try:
                    sc = SupplyChain(company_name=vendor)
                    db.add(sc)
                    db.flush()
                    existing_companies.add(vendor)
                    stats["supply_added"] += 1
                except Exception:
                    db.rollback()

            # 第2组供应链厂商 → 拼接到备注
            v2_name = row_data.pop("_vendor2_name", "")
            if v2_name and v2_name.strip():
                parts = [v2_name.strip()]
                for k2 in ("_vendor2_product", "_vendor2_version", "_vendor2_source"):
                    v2 = row_data.pop(k2, "")
                    if v2 and v2.strip():
                        parts.append(v2.strip())
                v2_text = " / ".join(parts)
                existing_remark = row_data.get("remark", "") or (getattr(existing, "remark", None) if existing else "")
                if existing_remark:
                    row_data["remark"] = str(existing_remark) + " | 供应商2: " + v2_text
                else:
                    row_data["remark"] = "供应商2: " + v2_text
            else:
                for k2 in ("_vendor2_name", "_vendor2_product", "_vendor2_version", "_vendor2_source"):
                    row_data.pop(k2, None)

        except Exception:
            db.rollback()
            stats["errors"] += 1

    db.commit()
    parts = [f"导入完成（{fmt_name}格式）：共 {stats['total']} 条"]
    if stats['created']: parts.append(f"新建 {stats['created']}")
    if stats['updated']: parts.append(f"覆盖 {stats['updated']}")
    if stats['supplemented']: parts.append(f"补充 {stats['supplemented']}")
    if stats['skipped']: parts.append(f"跳过 {stats['skipped']}")
    if stats['errors']: parts.append(f"失败 {stats['errors']}")
    if stats['dept_matched']: parts.append(f"匹配部门 {stats['dept_matched']}")
    if stats['users_created']: parts.append(f"创建账号 {stats['users_created']}")
    if stats['supply_added']: parts.append(f"新增供应链 {stats['supply_added']}")
    stats["message"] = "，".join(parts)
    return stats


# ── Excel 导出 ──

@router.get("/export")
def export_excel(db: Session = Depends(get_db), _=Depends(require_admin)):
    items = db.query(InfoSystem).order_by(InfoSystem.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "信息系统"
    # 中文表头映射（按逻辑分组排序）
    export_cols = [
        ("asset_id", "资产ID"),
        ("system_name", "系统名称"),
        ("system_type", "资产类型"),
        ("sub_type", "信息系统类型"),
        ("ip_address", "IP地址"),
        ("domain", "域名/URL"),
        ("org_name", "单位名称"),
        ("dept_name", "运维单位"),
        ("contact", "联系人"),
        ("contact_phone", "联系电话"),
        ("fill_type", "填报类型"),
        # 归属信息
        ("dept_id", "所属部门ID"),
        ("manager_name", "管理员姓名"),
        ("manager_gh", "管理员工号"),
        ("owner_name", "负责人姓名"),
        ("owner_gh", "负责人工号"),
        # 等级保护
        ("djdj_status", "等保状态"),
        ("djdj_sys_name", "等保系统名称"),
        ("djdj_no", "等保编号"),
        ("djdj_level", "等保等级"),
        ("djdj_date", "等保日期"),
        ("djdj_org", "测评单位"),
        # ICP
        ("icp_no", "ICP备案号"),
        ("icp_date", "ICP备案日期"),
        # 供应链
        ("vendor_name", "开发厂商"),
        ("product_name", "产品名称"),
        ("product_version", "版本号"),
        ("source_type", "来源"),
        ("vendor_contact", "厂商联系人"),
        ("vendor_phone", "厂商电话"),
        ("ops_contact", "运维联系人"),
        ("ops_phone", "运维电话"),
        # 其他
        ("has_website", "是否有网站"),
        ("internal_url", "内网URL"),
        ("ip_range", "IP地址段"),
        ("entry_url", "入口地址"),
        ("url_status", "验证状态"),
        ("remark", "备注"),
    ]
    headers = [h for _, h in export_cols]
    fields = [f for f, _ in export_cols]
    ws.append(headers)
    date_fmt_cols = {"djdj_date", "icp_date"}
    for r in items:
        row = []
        for f in fields:
            v = getattr(r, f, None)
            if v is None:
                row.append("")
            elif f in date_fmt_cols and hasattr(v, "strftime"):
                row.append(v.strftime("%Y-%m-%d"))
            else:
                row.append(v)
        ws.append(row)
    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = min(len(str(headers[col_idx - 1])) * 2, 30)
        for row_idx in range(2, min(ws.max_row + 1, 30)):
            cv = str(ws.cell(row_idx, col_idx).value or "")
            clen = sum(2 if ord(c) > 127 else 1 for c in cv)
            if clen > max_len:
                max_len = clen
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=info_systems.xlsx"})


# ── 等保记录 CRUD ──

@router.get("/djdj/image/{rec_id}")
def get_djdj_image(rec_id: int, db: Session = Depends(get_db)):
    rec = db.query(DjDjRecord).get(rec_id)
    if not rec or not rec.image_path:
        raise HTTPException(404, "无图片")
    filepath = f"uploads/{rec.image_path}"
    if not os.path.exists(filepath):
        raise HTTPException(404, "文件不存在")
    return FileResponse(filepath)


_DJ_SORTABLE = {"record_no","system_name","org_name","eval_org","level","record_date"}

@router.get("/djdj")
def list_djdj(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
              search: str = Query(""), sort_field: str = Query(""), sort_order: str = Query("desc"),
              db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(DjDjRecord)
    if not _is_admin(user):
        q = q.filter((DjDjRecord.claimed_by == user.id) | (DjDjRecord.claimed_by == None))
    if search:
        q = q.filter(DjDjRecord.system_name.like(f"%{search}%") | DjDjRecord.record_no.like(f"%{search}%"))
    total = q.count()
    if sort_field and sort_field in _DJ_SORTABLE:
        col = getattr(DjDjRecord, sort_field, None)
        q = q.order_by(col.desc()) if sort_order == "desc" and col else q.order_by(col.asc()) if col else q.order_by(DjDjRecord.id.desc())
    else:
        q = q.order_by(DjDjRecord.id.desc())
    items = q.offset((page - 1) * size).limit(size).all()
    result = []
    for r in items:
        d = {c.name: getattr(r, c.name) for c in r.__table__.columns}
        result.append(d)
    return {"items": result, "total": total}

@router.post("/djdj")
def create_djdj(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    data = dict(body)
    if "dept_name" in data and "eval_org" not in data:
        data["eval_org"] = data.pop("dept_name")
    elif "dept_name" in data:
        data.pop("dept_name")
    data["created_by"] = user.id
    rec = DjDjRecord(**{k: v for k, v in data.items() if hasattr(DjDjRecord, k)})
    db.add(rec); db.commit(); db.refresh(rec)
    return {"id": rec.id, "message": "已创建"}

@router.put("/djdj/{rec_id}")
def update_djdj(rec_id: int, body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(DjDjRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能编辑自己创建的数据")
    data = dict(body)
    if "dept_name" in data and "eval_org" not in data:
        data["eval_org"] = data.pop("dept_name")
    elif "dept_name" in data:
        data.pop("dept_name")
    try:
        for k, v in data.items():
            if hasattr(rec, k) and k != "id" and k not in _READONLY_COLS:
                setattr(rec, k, v)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"保存失败：{e}")
    return {"message": "已更新"}

@router.delete("/djdj/{rec_id}")
def delete_djdj(rec_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(DjDjRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能删除自己创建的数据")
    db.delete(rec); db.commit()
    return {"message": "已删除"}

@router.post("/djdj/batch-delete")
def batch_delete_djdj(body: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    ids = body.get("ids", [])
    if not ids: raise HTTPException(400, "请选择记录")
    db.query(DjDjRecord).filter(DjDjRecord.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {len(ids)} 条"}

@router.post("/djdj/upload-image/{rec_id}")
def upload_djdj_image(rec_id: int, file: UploadFile, db: Session = Depends(get_db), _=Depends(require_admin)):
    rec = db.query(DjDjRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    import os, shutil
    upload_dir = "uploads/djdj"
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or ".jpg")[1] or ".jpg"
    name = f"{rec_id}_{uuid.uuid4().hex[:8]}{ext}"
    filepath = f"{upload_dir}/{name}"
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    rec.image_path = f"djdj/{name}"
    db.commit()
    return {"message": "已上传", "image_path": rec.image_path}

@router.delete("/djdj/delete-image/{rec_id}")
def delete_djdj_image(rec_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    rec = db.query(DjDjRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if rec.image_path:
        import os
        filepath = f"uploads/{rec.image_path}"
        if os.path.exists(filepath):
            os.remove(filepath)
    rec.image_path = None
    db.commit()
    return {"message": "图片已删除"}

from datetime import date as date_type

def _find_data_start(ws):
    """跳过说明行和表头行，找到数据起始行。"""
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            s = str(val).strip()
            if s.isdigit() and len(s) >= 4:
                return row_idx
    return 3

def _cell_str(val) -> str:
    """安全转换单元格值为字符串。"""
    if val is None: return ""
    if isinstance(val, (int, float)): return str(int(val))
    if isinstance(val, date_type): return val.strftime("%Y-%m-%d")
    return str(val)


@router.post("/djdj/import")
def import_djdj(file: UploadFile, db: Session = Depends(get_db), _=Depends(require_admin)):
    try:
        data = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        imported = 0
        skipped = 0
        dupes = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = [_cell_str(v) for v in row]
            org = vals[1] if len(vals) > 1 else ""
            level = vals[2] if len(vals) > 2 else ""
            sys_name = vals[3] if len(vals) > 3 else ""
            rn = vals[4] if len(vals) > 4 else ""
            rd = vals[5] if len(vals) > 5 and vals[5] else None
            if not rn or not sys_name:
                continue
            exist = db.execute(text("SELECT id,system_name,org_name FROM djdj_records WHERE record_no=:r"), {"r": rn}).fetchone()
            if exist:
                skipped += 1
                dupes.append({"备案编号": rn, "导入系统名": sys_name, "已存在系统名": exist.system_name, "导入单位": org, "已存在单位": exist.org_name})
                continue
            db.add(DjDjRecord(record_no=rn, system_name=sys_name, org_name=org, level=level, record_date=rd))
            db.flush()
            imported += 1
        db.commit()
        return {"message": f"导入完成：新增 {imported} 条，跳过重复 {skipped} 条", "duplicates": dupes[:20]}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"导入失败：{e}")

@router.get("/djdj/export")
def export_djdj(db: Session = Depends(get_db), _=Depends(require_admin)):
    items = db.query(DjDjRecord).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "等保备案"
    ws.append(["备案编号", "系统名称", "备案单位", "等级", "备案日期", "测评单位"])
    for r in items:
        ws.append([r.record_no, r.system_name, r.org_name, r.level,
                    str(r.record_date) if r.record_date else "", r.eval_org])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=djdj_export.xlsx"})


# ── ICP记录 CRUD ──

_ICP_SORTABLE = {"icp_no","org_name","domain","record_date"}

@router.get("/icp")
def list_icp(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
             search: str = Query(""), sort_field: str = Query(""), sort_order: str = Query("desc"),
             db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(IcpRecord)
    if not _is_admin(user):
        q = q.filter((IcpRecord.claimed_by == user.id) | (IcpRecord.claimed_by == None))
    if search:
        q = q.filter(IcpRecord.icp_no.like(f"%{search}%") | IcpRecord.org_name.like(f"%{search}%"))
    total = q.count()
    if sort_field and sort_field in _ICP_SORTABLE:
        col = getattr(IcpRecord, sort_field, None)
        q = q.order_by(col.desc()) if sort_order == "desc" and col else q.order_by(col.asc()) if col else q.order_by(IcpRecord.id.desc())
    else:
        q = q.order_by(IcpRecord.id.desc())
    items = q.offset((page - 1) * size).limit(size).all()
    return {"items": [{c.name: getattr(r, c.name) for c in r.__table__.columns} for r in items], "total": total}

@router.post("/icp")
def create_icp(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    data = {k: v for k, v in body.items() if hasattr(IcpRecord, k)}
    data["created_by"] = user.id
    rec = IcpRecord(**data)
    db.add(rec); db.commit(); db.refresh(rec)
    return {"id": rec.id, "message": "已创建"}

@router.put("/icp/{rec_id}")
def update_icp(rec_id: int, body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(IcpRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能编辑自己创建的数据")
    try:
        for k, v in body.items():
            if hasattr(rec, k) and k != "id" and k not in _READONLY_COLS:
                setattr(rec, k, v)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"保存失败：{e}")
    return {"message": "已更新"}

@router.delete("/icp/{rec_id}")
def delete_icp(rec_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(IcpRecord).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能删除自己创建的数据")
    db.delete(rec); db.commit()
    return {"message": "已删除"}


# ── 数据同步 ──

@router.post("/sync-from-platform")
def sync_from_platform(db: Session = Depends(get_db), _=Depends(require_admin)):
    """域名清洗 + ZDNS 匹配同步（手动触发，写入扫描日志）。"""
    from app.services.scheduler_service import run_info_system_sync_manual
    try:
        result = run_info_system_sync_manual()
        return result
    except Exception as e:
        raise HTTPException(500, f"同步失败：{e}")


# ── 等保关联查询 ──

@router.get("/djdj-search")
def search_djdj(q: str = Query(""), db: Session = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(DjDjRecord).filter(DjDjRecord.system_name.contains(q)).limit(20).all()
    return {"items": [{"id": r.id, "system_name": r.system_name, "record_no": r.record_no,
                        "level": r.level, "record_date": str(r.record_date) if r.record_date else None,
                        "org_name": r.org_name} for r in items]}


# ── 供应链 CRUD ──

_SC_SORTABLE = {"company_name","credit_code","company_type","importance","security_contact","security_phone"}

@router.get("/supply-chain")
def list_supply_chain(page: int = Query(1, ge=1), size: int = Query(20, ge=1, le=100),
                      search: str = Query(""), sort_field: str = Query(""), sort_order: str = Query("desc"),
                      db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(SupplyChain)
    if not _is_admin(user):
        q = q.filter((SupplyChain.claimed_by == user.id) | (SupplyChain.claimed_by == None))
    if search:
        q = q.filter(SupplyChain.company_name.contains(search))
    total = q.count()
    if sort_field and sort_field in _SC_SORTABLE:
        col = getattr(SupplyChain, sort_field, None)
        q = q.order_by(col.desc()) if sort_order == "desc" and col else q.order_by(col.asc()) if col else q.order_by(SupplyChain.id.desc())
    else:
        q = q.order_by(SupplyChain.id.desc())
    items = q.offset((page - 1) * size).limit(size).all()
    return {"items": [{c.name: getattr(r, c.name) for c in r.__table__.columns} for r in items], "total": total}

@router.post("/supply-chain")
def create_supply_chain(body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    data = {k: v for k, v in body.items() if hasattr(SupplyChain, k)}
    data["created_by"] = user.id
    rec = SupplyChain(**data)
    db.add(rec); db.commit(); db.refresh(rec)
    return {"id": rec.id, "message": "已创建"}

@router.get("/supply-chain/names")
def list_supply_chain_names(db: Session = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(SupplyChain.company_name).order_by(SupplyChain.company_name).all()
    return {"items": [r[0] for r in items]}

# ── 供应链导入导出 ──

# Excel 行业名称到系统选项的映射（Excel 中使用中文顿号分隔，系统使用斜杠）
_INDUSTRY_MAP = {
    "信息运输、软件和信息技术服务业": "信息运输/软件和信息技术服务业",
    "文化、体育和娱乐业": "文化/体育和娱乐业",
    "交通运输、仓储和邮政业": "交通运输/仓储和邮政业",
    "电力、热力、燃气及水生产和供应业": "电力/热力/燃气及水生产和供应业",
    "农、林、牧、渔业": "农/林/牧/渔业",
    "公共管理、社会保障和社会组织": "公共管理/社会保障和社会组织",
    "卫生和社会工作": "卫生和社会工作",
    "居民服务、修理和其他服务业": "居民服务/修理和其他服务业",
    "水利、环境和公共": "水利/环境和公共",
    "科学研究和技术服务业": "科学研究和技术服务业",
    "租聘和商务": "租聘和商务",
    "批发和零售业": "批发和零售业",
    "住宿和餐饮业": "住宿和餐饮业",
    "制造业": "制造业",
    "建筑业": "建筑业",
    "采矿业": "采矿业",
}


def _normalize_industry(val: str) -> str:
    """将 Excel 中的行业名称转换为系统选项格式。"""
    parts = [p.strip() for p in val.split(",") if p.strip()]
    return ",".join(_INDUSTRY_MAP.get(p, p) for p in parts)


@router.post("/supply-chain/import")
def import_supply_chain(file: UploadFile, db: Session = Depends(get_db), _=Depends(require_admin)):
    """从信息系统管理 Excel 文件中导入供应链单位信息（去重）。"""
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    # 查询已有单位名称用于去重
    existing = {r[0] for r in db.query(SupplyChain.company_name).all()}
    imported, skipped = 0, 0
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [_cell_str(v) for v in row]
        if len(vals) <= 51:
            continue
        name = vals[51].strip()
        if not name:
            continue
        if name in seen or name in existing:
            skipped += 1
            continue
        seen.add(name)
        # 行业：合并主行业(col 59) + 其他(col 60)
        industry_raw = vals[59] if len(vals) > 59 else ""
        industry_other = vals[60] if len(vals) > 60 else ""
        if industry_raw and industry_other:
            industry_raw = industry_raw + "," + industry_other
        industry = _normalize_industry(industry_raw) if industry_raw else ""
        # 服务类型：合并 col 61 + col 62
        service_raw = vals[61] if len(vals) > 61 else ""
        service_other = vals[62] if len(vals) > 62 else ""
        if service_raw and service_other and service_other != "无":
            service_raw = service_raw + "," + service_other
        rec = SupplyChain(
            company_name=name,
            credit_code=vals[52] if len(vals) > 52 else "",
            address=vals[53] if len(vals) > 53 else "",
            security_dept=vals[54] if len(vals) > 54 else "",
            security_contact=vals[55] if len(vals) > 55 else "",
            security_phone=vals[56] if len(vals) > 56 else "",
            company_type=vals[57] if len(vals) > 57 else "",
            has_foreign_capital=vals[58] if len(vals) > 58 else "",
            industry=industry,
            service_type=service_raw if service_raw else "",
            importance=vals[63] if len(vals) > 63 else "",
            url_ip_range=vals[64] if len(vals) > 64 else "",
            data_level=vals[65] if len(vals) > 65 else "",
            data_location=vals[66] if len(vals) > 66 else "",
            data_storage=vals[67] if len(vals) > 67 else "",
            db_type=vals[68] if len(vals) > 68 else "",
        )
        db.add(rec)
        imported += 1
    db.commit()
    return {"message": f"导入完成：新增 {imported} 条，跳过重复 {skipped} 条"}


@router.get("/supply-chain/export")
def export_supply_chain(db: Session = Depends(get_db), _=Depends(require_admin)):
    items = db.query(SupplyChain).order_by(SupplyChain.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应链信息"
    ws.append(["单位名称", "信用代码", "注册地址", "安全责任部门", "联系人", "联系电话",
               "单位类型", "境外资本", "服务行业", "服务类型", "重要程度",
               "URL/IP地址段", "数据最高级别", "存储位置", "存储方式", "数据库类型", "备注"])
    for r in items:
        ws.append([r.company_name, r.credit_code, r.address, r.security_dept,
                   r.security_contact, r.security_phone, r.company_type,
                   r.has_foreign_capital, r.industry, r.service_type, r.importance,
                   r.url_ip_range, r.data_level, r.data_location, r.data_storage,
                   r.db_type, r.remark or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=supply_chain_export.xlsx"}
    )

@router.post("/supply-chain/batch-delete")
def batch_delete_supply_chain(body: dict, db: Session = Depends(get_db), _=Depends(require_admin)):
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "请选择记录")
    db.query(SupplyChain).filter(SupplyChain.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {len(ids)} 条"}


@router.put("/supply-chain/{rec_id}")
def update_supply_chain(rec_id: int, body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(SupplyChain).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能编辑自己创建的数据")
    try:
        for k, v in body.items():
            if hasattr(rec, k) and k != "id" and k not in _READONLY_COLS:
                setattr(rec, k, v)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"保存失败：{e}")
    return {"message": "已更新"}

@router.delete("/supply-chain/{rec_id}")
def delete_supply_chain(rec_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    rec = db.query(SupplyChain).get(rec_id)
    if not rec: raise HTTPException(404, "不存在")
    if not _is_admin(user) and rec.created_by != user.id:
        raise HTTPException(403, "只能删除自己创建的数据")
    db.delete(rec); db.commit()
    return {"message": "已删除"}


@router.post("/icp/import")
def import_icp(file: UploadFile, db: Session = Depends(get_db), _=Depends(require_admin)):
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    start = _find_data_start(ws)
    imported = 0
    db.query(IcpRecord).delete()
    for row in ws.iter_rows(min_row=start, values_only=True):
        vals = [_cell_str(v) for v in row]
        if not vals[0] or not vals[0].isdigit(): continue
        rec = IcpRecord(
            icp_no=vals[4] if len(vals) > 4 else "",
            org_name=vals[2] if len(vals) > 2 else "",
            domain=vals[3] if len(vals) > 3 else "",
            record_date=vals[5] if len(vals) > 5 and vals[5] else None,
        )
        db.add(rec); imported += 1
    db.commit()
    return {"message": f"导入完成，共 {imported} 条"}

@router.get("/icp/export")
def export_icp(db: Session = Depends(get_db), _=Depends(require_admin)):
    items = db.query(IcpRecord).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ICP备案"
    ws.append(["ICP备案编号", "备案主体", "备案域名", "备案日期"])
    for r in items:
        ws.append([r.icp_no, r.org_name, r.domain, str(r.record_date) if r.record_date else ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=icp_export.xlsx"})
